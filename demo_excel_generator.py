#!/usr/bin/env python3
"""
Demo Excel Signal Generator for GenX FX Trading System
This script creates sample forex trading signals and exports them to Excel/CSV
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import os

class ForexSignalGenerator:
    """
    A class to generate and export demo forex trading signals to various formats.
    """
    def __init__(self):
        """
        Initializes the ForexSignalGenerator, setting up pairs, timeframes, and output directory.
        """
        self.pairs = ['EURUSD', 'GBPUSD', 'USDJPY', 'USDCHF', 'AUDUSD', 'USDCAD', 'NZDUSD',
                      'XAUUSD', 'XAUEUR', 'XAUGBP', 'XAUAUD', 'XAUCAD', 'XAUCHF']
        self.timeframes = ['M15', 'H1', 'H4', 'D1']
        self.signal_output_dir = 'signal_output'
        
        # Create output directory
        os.makedirs(self.signal_output_dir, exist_ok=True)
        
    def generate_price_data(self, symbol: str) -> float:
        """
        Generates realistic price data for a given symbol.

        Args:
            symbol (str): The financial instrument symbol (e.g., 'EURUSD').

        Returns:
            float: A randomly generated price for the symbol.
        """
        base_prices = {
            'EURUSD': 1.10500, 'GBPUSD': 1.27000, 'USDJPY': 149.50,
            'USDCHF': 0.88200, 'AUDUSD': 0.65800, 'USDCAD': 1.36500, 'NZDUSD': 0.58900,
            'XAUUSD': 2650.00, 'XAUEUR': 2420.00, 'XAUGBP': 2100.00, 
            'XAUAUD': 4050.00, 'XAUCAD': 3620.00, 'XAUCHF': 2340.00
        }
        
        base_price = base_prices.get(symbol, 1.0000)
        
        # Add some realistic variation
        variation = random.uniform(-0.005, 0.005)  # 0.5% variation
        current_price = base_price * (1 + variation)
        
        return round(current_price, 5)
    
    def generate_signal_data(self, num_signals: int = 10) -> pd.DataFrame:
        """
        Generates a DataFrame of sample trading signals.

        Args:
            num_signals (int, optional): The number of signals to generate. Defaults to 10.

        Returns:
            pd.DataFrame: A DataFrame containing the generated signal data.
        """
        signals = []
        
        for i in range(num_signals):
            symbol = random.choice(self.pairs)
            current_price = self.generate_price_data(symbol)
            
            # Generate signal
            signal_type = random.choice(['BUY', 'SELL'])
            
            # Calculate realistic stop loss and take profit
            pip_value = 0.0001 if 'JPY' not in symbol else 0.01
            
            if signal_type == 'BUY':
                entry_price = current_price + random.uniform(0, 10) * pip_value
                stop_loss = entry_price - random.uniform(20, 50) * pip_value
                take_profit = entry_price + random.uniform(40, 100) * pip_value
            else:
                entry_price = current_price - random.uniform(0, 10) * pip_value
                stop_loss = entry_price + random.uniform(20, 50) * pip_value
                take_profit = entry_price - random.uniform(40, 100) * pip_value
            
            # Calculate lot size (risk-based)
            risk_reward = abs(take_profit - entry_price) / abs(entry_price - stop_loss)
            lot_size = round(random.uniform(0.01, 0.10), 2)
            confidence = round(random.uniform(0.75, 0.95), 2)
            
            signal = {
                'Timestamp': datetime.now() + timedelta(minutes=random.randint(-30, 30)),
                'Symbol': symbol,
                'Signal': signal_type,
                'Entry_Price': round(entry_price, 5),
                'Stop_Loss': round(stop_loss, 5),
                'Take_Profit': round(take_profit, 5),
                'Lot_Size': lot_size,
                'Confidence': confidence,
                'Risk_Reward': round(risk_reward, 2),
                'Magic_Number': 123450 + i,
                'Timeframe': random.choice(self.timeframes),
                'Status': random.choice(['Active', 'Pending', 'Filled']),
                'Comment': f'GenX_{signal_type}_{symbol}'
            }
            
            signals.append(signal)
        
        return pd.DataFrame(signals)
    
    def create_excel_dashboard(self, df: pd.DataFrame) -> str:
        """
        Creates a professional Excel dashboard from the signal data.

        Args:
            df (pd.DataFrame): The DataFrame containing signal data.

        Returns:
            str: The file path of the created Excel dashboard.
        """
        file_path = os.path.join(self.signal_output_dir, 'genx_signals.xlsx')
        
        # Create workbook and worksheets
        wb = openpyxl.Workbook()
        
        # Remove default worksheet
        wb.remove(wb.active)
        
        # Create Active Signals sheet
        ws_active = wb.create_sheet('Active Signals')
        ws_summary = wb.create_sheet('Summary Dashboard')
        ws_history = wb.create_sheet('Signal History')
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        buy_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
        sell_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Active Signals Sheet
        active_signals = df[df['Status'] == 'Active'].copy()
        
        # Write headers
        headers = ['Timestamp', 'Symbol', 'Signal', 'Entry Price', 'Stop Loss', 
                  'Take Profit', 'Lot Size', 'Confidence', 'Risk/Reward', 'Status']
        
        for col, header in enumerate(headers, 1):
            cell = ws_active.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data
        for row_idx, (_, signal) in enumerate(active_signals.iterrows(), 2):
            ws_active.cell(row=row_idx, column=1, value=signal['Timestamp'].strftime('%Y-%m-%d %H:%M'))
            ws_active.cell(row=row_idx, column=2, value=signal['Symbol'])
            ws_active.cell(row=row_idx, column=3, value=signal['Signal'])
            ws_active.cell(row=row_idx, column=4, value=signal['Entry_Price'])
            ws_active.cell(row=row_idx, column=5, value=signal['Stop_Loss'])
            ws_active.cell(row=row_idx, column=6, value=signal['Take_Profit'])
            ws_active.cell(row=row_idx, column=7, value=signal['Lot_Size'])
            ws_active.cell(row=row_idx, column=8, value=f"{signal['Confidence']:.1%}")
            ws_active.cell(row=row_idx, column=9, value=signal['Risk_Reward'])
            ws_active.cell(row=row_idx, column=10, value=signal['Status'])
            
            # Color coding for BUY/SELL
            signal_cell = ws_active.cell(row=row_idx, column=3)
            if signal['Signal'] == 'BUY':
                signal_cell.fill = buy_fill
            else:
                signal_cell.fill = sell_fill
            
            # Apply borders
            for col in range(1, 11):
                ws_active.cell(row=row_idx, column=col).border = border
        
        # Auto-adjust column widths
        for col in range(1, 11):
            ws_active.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Summary Dashboard
        ws_summary.cell(row=1, column=1, value='GenX FX Trading Dashboard')
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        summary_data = [
            ['Total Signals', len(df)],
            ['Active Signals', len(df[df['Status'] == 'Active'])],
            ['BUY Signals', len(df[df['Signal'] == 'BUY'])],
            ['SELL Signals', len(df[df['Signal'] == 'SELL'])],
            ['Average Confidence', f"{df['Confidence'].mean():.1%}"],
            ['Average Risk/Reward', f"{df['Risk_Reward'].mean():.2f}"],
            ['Last Update', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Signal History (all signals)
        for col, header in enumerate(['Timestamp', 'Symbol', 'Signal', 'Entry', 'SL', 'TP', 'Lots', 'Confidence', 'R:R', 'Status'], 1):
            cell = ws_history.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, signal) in enumerate(df.iterrows(), 2):
            ws_history.cell(row=row_idx, column=1, value=signal['Timestamp'].strftime('%Y-%m-%d %H:%M'))
            ws_history.cell(row=row_idx, column=2, value=signal['Symbol'])
            ws_history.cell(row=row_idx, column=3, value=signal['Signal'])
            ws_history.cell(row=row_idx, column=4, value=signal['Entry_Price'])
            ws_history.cell(row=row_idx, column=5, value=signal['Stop_Loss'])
            ws_history.cell(row=row_idx, column=6, value=signal['Take_Profit'])
            ws_history.cell(row=row_idx, column=7, value=signal['Lot_Size'])
            ws_history.cell(row=row_idx, column=8, value=signal['Confidence'])
            ws_history.cell(row=row_idx, column=9, value=signal['Risk_Reward'])
            ws_history.cell(row=row_idx, column=10, value=signal['Status'])
        
        # Save workbook
        wb.save(file_path)
        print(f"✅ Excel dashboard created: {file_path}")
        return file_path
    
    def create_mt4_csv(self, df: pd.DataFrame) -> str:
        """
        Creates a MT4-compatible CSV file from the signal data.

        Args:
            df (pd.DataFrame): The DataFrame containing signal data.

        Returns:
            str: The file path of the created MT4 CSV file.
        """
        file_path = os.path.join(self.signal_output_dir, 'MT4_Signals.csv')
        
        mt4_data = df[['Magic_Number', 'Symbol', 'Signal', 'Entry_Price', 
                       'Stop_Loss', 'Take_Profit', 'Lot_Size', 'Timestamp']].copy()
        
        mt4_data.columns = ['Magic', 'Symbol', 'Signal', 'EntryPrice', 
                           'StopLoss', 'TakeProfit', 'LotSize', 'Timestamp']
        
        mt4_data['Timestamp'] = mt4_data['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        mt4_data.to_csv(file_path, index=False)
        print(f"✅ MT4 CSV created: {file_path}")
        return file_path
    
    def create_mt5_csv(self, df: pd.DataFrame) -> str:
        """
        Creates a MT5-compatible CSV file from the signal data.

        Args:
            df (pd.DataFrame): The DataFrame containing signal data.

        Returns:
            str: The file path of the created MT5 CSV file.
        """
        file_path = os.path.join(self.signal_output_dir, 'MT5_Signals.csv')
        
        mt5_data = df.copy()
        mt5_data['Volume'] = mt5_data['Lot_Size']
        mt5_data['Expiry'] = (mt5_data['Timestamp'] + timedelta(hours=4)).dt.strftime('%Y-%m-%d %H:%M:%S')
        mt5_data['Timestamp'] = mt5_data['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        mt5_columns = ['Magic_Number', 'Symbol', 'Signal', 'Entry_Price', 'Stop_Loss', 
                      'Take_Profit', 'Volume', 'Confidence', 'Risk_Reward', 'Expiry', 'Comment']
        
        mt5_data[mt5_columns].to_csv(file_path, index=False)
        print(f"✅ MT5 CSV created: {file_path}")
        return file_path
    
    def create_json_output(self, df: pd.DataFrame) -> str:
        """
        Creates a JSON output file from the signal data.

        Args:
            df (pd.DataFrame): The DataFrame containing signal data.

        Returns:
            str: The file path of the created JSON file.
        """
        file_path = os.path.join(self.signal_output_dir, 'genx_signals.json')
        
        # Convert timestamps to strings for JSON serialization
        df_json = df.copy()
        df_json['Timestamp'] = df_json['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        output = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_signals': len(df),
            'active_signals': len(df[df['Status'] == 'Active']),
            'signals': df_json.to_dict('records')
        }
        
        import json
        with open(file_path, 'w') as f:
            json.dump(output, f, indent=2, default=str)
        
        print(f"✅ JSON output created: {file_path}")
        return file_path
    
    def run_demo(self, num_signals: int = 15) -> tuple:
        """
        Runs the complete demo, generating signals and creating all output files.

        Args:
            num_signals (int, optional): The number of signals to generate. Defaults to 15.

        Returns:
            tuple: A tuple containing the file paths of the created Excel, MT4, MT5, and JSON files.
        """
        print("🚀 GenX FX Trading System - Excel Demo")
        print("=" * 50)
        
        # Generate signals
        print(f"📊 Generating {num_signals} sample forex signals...")
        df = self.generate_signal_data(num_signals)
        
        # Create outputs
        excel_file = self.create_excel_dashboard(df)
        mt4_file = self.create_mt4_csv(df)
        mt5_file = self.create_mt5_csv(df)
        json_file = self.create_json_output(df)
        
        # Print summary
        print("\n📈 Signal Summary:")
        print(f"   • Total Signals: {len(df)}")
        print(f"   • BUY Signals: {len(df[df['Signal'] == 'BUY'])}")
        print(f"   • SELL Signals: {len(df[df['Signal'] == 'SELL'])}")
        print(f"   • Average Confidence: {df['Confidence'].mean():.1%}")
        print(f"   • Average Risk/Reward: {df['Risk_Reward'].mean():.2f}")
        
        print(f"\n📁 Output Files:")
        print(f"   📊 Excel Dashboard: {excel_file}")
        print(f"   📈 MT4 Signals: {mt4_file}")
        print(f"   📈 MT5 Signals: {mt5_file}")
        print(f"   🔗 JSON API: {json_file}")
        
        print(f"\n✨ Demo completed successfully!")
        print(f"📂 Check the '{self.signal_output_dir}' directory for all output files.")
        
        return excel_file, mt4_file, mt5_file, json_file

if __name__ == "__main__":
    generator = ForexSignalGenerator()
    generator.run_demo()
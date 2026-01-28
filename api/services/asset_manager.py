"""
Asset Management Service for Google Sheets/Excel integration.
Provides centralized portfolio tracking, risk management, and trade logging.
"""

import asyncio
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, asdict
import pandas as pd
import gspread
from gspread_asyncio import AsyncioGspreadClientManager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from google.oauth2.service_account import Credentials
import json
import os
from ..config.settings import get_settings

settings = get_settings()
logger = logging.getLogger(__name__)


@dataclass
class Position:
    """
    Represents an open trading position.

    Attributes:
        ticket (str): The unique ticket or ID for the position.
        symbol (str): The trading symbol (e.g., 'EUR_USD').
        type (str): The type of position, either "BUY" or "SELL".
        lots (float): The volume of the position in lots.
        open_price (float): The price at which the position was opened.
        current_price (float): The current market price of the symbol.
        stop_loss (Optional[float]): The stop-loss price, if set.
        take_profit (Optional[float]): The take-profit price, if set.
        open_time (datetime): The timestamp when the position was opened.
        current_pnl (float): The current profit or loss, usually provided by the broker.
        commission (float): Any commission paid for the trade.
        swap (float): Any swap fees incurred.
    """

    ticket: str
    symbol: str
    type: str  # "BUY" or "SELL"
    lots: float
    open_price: float
    current_price: float
    stop_loss: Optional[float]
    take_profit: Optional[float]
    open_time: datetime
    current_pnl: float
    commission: float = 0.0
    swap: float = 0.0

    @property
    def unrealized_pnl(self) -> float:
        """
        Calculates the unrealized Profit & Loss (P&L) for the position.

        Note: This is a simplified calculation and may differ from broker calculations.
        It assumes a standard contract size.

        Returns:
            float: The calculated unrealized P&L.
        """
        if self.type == "BUY":
            return (self.current_price - self.open_price) * self.lots * 100000
        else:
            return (self.open_price - self.current_price) * self.lots * 100000


@dataclass
class ClosedTrade:
    """
    Represents a trade that has been closed.

    Attributes:
        ticket (str): The unique ticket or ID for the trade.
        symbol (str): The trading symbol.
        type (str): The type of trade, "BUY" or "SELL".
        lots (float): The volume of the trade in lots.
        open_price (float): The price at which the trade was opened.
        close_price (float): The price at which the trade was closed.
        stop_loss (Optional[float]): The stop-loss price that was set.
        take_profit (Optional[float]): The take-profit price that was set.
        open_time (datetime): The timestamp when the trade was opened.
        close_time (datetime): The timestamp when the trade was closed.
        realized_pnl (float): The final profit or loss from the trade.
        commission (float): Total commission paid.
        swap (float): Total swap fees incurred.
    """

    ticket: str
    symbol: str
    type: str
    lots: float
    open_price: float
    close_price: float
    stop_loss: Optional[float]
    take_profit: Optional[float]
    open_time: datetime
    close_time: datetime
    realized_pnl: float
    commission: float = 0.0
    swap: float = 0.0

    @property
    def duration_hours(self) -> float:
        """
        Calculates the duration of the trade in hours.

        Returns:
            float: The total duration of the trade in hours.
        """
        return (self.close_time - self.open_time).total_seconds() / 3600


@dataclass
class AccountSummary:
    """
    Holds a summary of the trading account's status and performance.

    Attributes:
        balance (float): The current account balance.
        equity (float): The current account equity (balance + unrealized P&L).
        margin (float): The amount of margin currently in use.
        free_margin (float): The margin available for new trades.
        margin_level (float): The margin level percentage.
        total_open_positions (int): The number of currently open positions.
        total_unrealized_pnl (float): The total P&L of all open positions.
        daily_pnl (float): The realized P&L for the current day.
        weekly_pnl (float): The realized P&L for the current week.
        monthly_pnl (float): The realized P&L for the current month.
        max_drawdown (float): The maximum drawdown percentage observed.
        win_rate (float): The win rate of closed trades.
        profit_factor (float): The ratio of gross profit to gross loss.
        last_updated (datetime): The timestamp of the last summary update.
    """

    balance: float
    equity: float
    margin: float
    free_margin: float
    margin_level: float
    total_open_positions: int
    total_unrealized_pnl: float
    daily_pnl: float
    weekly_pnl: float
    monthly_pnl: float
    max_drawdown: float
    win_rate: float
    profit_factor: float
    last_updated: datetime


@dataclass
class RiskParameters:
    """
    Defines the risk management parameters for the trading strategy.

    Attributes:
        max_risk_per_trade (float): Max risk per trade as a percentage of equity (e.g., 0.02 for 2%).
        max_daily_drawdown (float): Max daily drawdown limit as a percentage.
        max_correlation (float): Maximum allowed correlation between open positions.
        max_exposure_per_currency (float): Max exposure per currency as a percentage of equity.
        max_lot_size (float): Maximum lot size allowed for a single trade.
        max_open_positions (int): Maximum number of concurrent open positions.
        instruments_to_trade (List[str]): A list of symbols that are allowed to be traded.
    """

    max_risk_per_trade: float = 0.02  # 2% of equity
    max_daily_drawdown: float = 0.05  # 5% daily drawdown limit
    max_correlation: float = 0.7  # Maximum correlation between positions
    max_exposure_per_currency: float = 0.1  # 10% exposure per currency
    max_lot_size: float = 1.0  # Maximum lot size per trade
    max_open_positions: int = 10  # Maximum number of open positions
    instruments_to_trade: Optional[List[str]] = None

    def __post_init__(self):
        """Sets default instruments if none are provided."""
        if self.instruments_to_trade is None:
            self.instruments_to_trade = [
                "EUR_USD",
                "GBP_USD",
                "USD_JPY",
                "USD_CHF",
                "AUD_USD",
                "USD_CAD",
                "NZD_USD",
                "EUR_GBP",
            ]


class GoogleSheetsManager:
    """
    Manages integration with Google Sheets for portfolio tracking and reporting.

    This class handles authentication, worksheet creation, formatting, and data updates.

    Attributes:
        credentials_path (str): Path to the Google service account credentials JSON file.
        spreadsheet_key (str): The key of the Google Sheet to interact with.
        client_manager (AsyncioGspreadClientManager): The asynchronous gspread client manager.
        worksheet_names (Dict[str, str]): A mapping of internal names to actual worksheet titles.
    """

    def __init__(self, credentials_path: str, spreadsheet_key: str):
        """
        Initializes the GoogleSheetsManager.

        Args:
            credentials_path (str): The path to the service account credentials file.
            spreadsheet_key (str): The key of the target Google Sheet.
        """
        self.credentials_path = credentials_path
        self.spreadsheet_key = spreadsheet_key
        self.client_manager = None
        self.worksheet_names = {
            "dashboard": "Dashboard",
            "open_positions": "Open Positions",
            "closed_trades": "Closed Trades",
            "risk_config": "Risk Configuration",
            "performance": "Performance Analytics",
        }

    async def initialize(self):
        """
        Initializes the asynchronous Google Sheets client.

        Sets up the client manager with the provided service account credentials.

        Raises:
            Exception: If the client fails to initialize.
        """
        try:

            def get_creds():
                return Credentials.from_service_account_file(
                    self.credentials_path,
                    scopes=[
                        "https://www.googleapis.com/auth/spreadsheets",
                        "https://www.googleapis.com/auth/drive",
                    ],
                )

            self.client_manager = AsyncioGspreadClientManager(get_creds)
            logger.info("Google Sheets client initialized successfully")

        except Exception as e:
            logger.error(f"Failed to initialize Google Sheets client: {e}")
            raise

    async def get_spreadsheet(self):
        """
        Authorizes the client and opens the spreadsheet by its key.

        Returns:
            gspread_asyncio.AsyncioGspreadSpreadsheet: The spreadsheet object.
        """
        client = await self.client_manager.authorize()
        return await client.open_by_key(self.spreadsheet_key)

    async def setup_worksheets(self):
        """
        Creates and formats all required worksheets if they don't already exist.

        This method ensures the spreadsheet has the necessary structure for reporting.

        Raises:
            Exception: If worksheet setup fails.
        """
        try:
            spreadsheet = await self.get_spreadsheet()

            # Create worksheets if they don't exist
            existing_sheets = [ws.title for ws in await spreadsheet.worksheets()]
            for sheet_name in self.worksheet_names.values():
                if sheet_name not in existing_sheets:
                    await spreadsheet.add_worksheet(
                        title=sheet_name, rows=1000, cols=20
                    )
                    logger.info(f"Created worksheet: {sheet_name}")

            # Format each worksheet
            await self._format_dashboard()
            await self._format_positions_sheet()
            await self._format_trades_sheet()
            await self._format_risk_config()
            await self._format_performance_sheet()

        except Exception as e:
            logger.error(f"Error setting up worksheets: {e}")
            raise

    async def _format_dashboard(self):
        """Formats the 'Dashboard' worksheet with headers and structure."""
        spreadsheet = await self.get_spreadsheet()
        worksheet = await spreadsheet.worksheet(self.worksheet_names["dashboard"])

        # Headers and structure
        headers = [
            ["Account Summary", "", "", ""],
            ["Metric", "Value", "Previous", "Change"],
            ["Balance", "", "", ""],
            ["Equity", "", "", ""],
            ["Free Margin", "", "", ""],
            ["Margin Level", "", "", ""],
            ["Open Positions", "", "", ""],
            ["Unrealized P&L", "", "", ""],
            ["Daily P&L", "", "", ""],
            ["Weekly P&L", "", "", ""],
            ["Monthly P&L", "", "", ""],
            ["", "", "", ""],
            ["Risk Metrics", "", "", ""],
            ["Max Drawdown", "", "", ""],
            ["Win Rate", "", "", ""],
            ["Profit Factor", "", "", ""],
            ["", "", "", ""],
            ["Last Updated", "", "", ""],
        ]

        await worksheet.update("A1:D18", headers)

        # Format headers
        await worksheet.format(
            "A1:D1",
            {
                "textFormat": {"bold": True, "fontSize": 14},
                "backgroundColor": {"red": 0.2, "green": 0.2, "blue": 0.8},
            },
        )

    async def _format_positions_sheet(self):
        """Formats the 'Open Positions' worksheet with headers."""
        spreadsheet = await self.get_spreadsheet()
        worksheet = await spreadsheet.worksheet(self.worksheet_names["open_positions"])

        headers = [
            "Ticket",
            "Symbol",
            "Type",
            "Lots",
            "Open Price",
            "Current Price",
            "Stop Loss",
            "Take Profit",
            "Open Time",
            "Unrealized P&L",
            "Commission",
            "Swap",
        ]

        await worksheet.update("A1:L1", [headers])
        await worksheet.format(
            "A1:L1",
            {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.8, "green": 0.2, "blue": 0.2},
            },
        )

    async def _format_trades_sheet(self):
        """Formats the 'Closed Trades' worksheet with headers."""
        spreadsheet = await self.get_spreadsheet()
        worksheet = await spreadsheet.worksheet(self.worksheet_names["closed_trades"])

        headers = [
            "Ticket",
            "Symbol",
            "Type",
            "Lots",
            "Open Price",
            "Close Price",
            "Stop Loss",
            "Take Profit",
            "Open Time",
            "Close Time",
            "Duration (h)",
            "Realized P&L",
            "Commission",
            "Swap",
        ]

        await worksheet.update("A1:N1", [headers])
        await worksheet.format(
            "A1:N1",
            {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.2, "green": 0.8, "blue": 0.2},
            },
        )

    async def _format_risk_config(self):
        """Formats the 'Risk Configuration' worksheet with default parameters."""
        spreadsheet = await self.get_spreadsheet()
        worksheet = await spreadsheet.worksheet(self.worksheet_names["risk_config"])

        config_data = [
            ["Risk Parameter", "Value", "Description"],
            ["Max Risk Per Trade", "0.02", "Maximum risk per trade (% of equity)"],
            ["Max Daily Drawdown", "0.05", "Maximum daily drawdown limit"],
            ["Max Correlation", "0.7", "Maximum correlation between positions"],
            ["Max Exposure Per Currency", "0.1", "Maximum exposure per currency"],
            ["Max Lot Size", "1.0", "Maximum lot size per trade"],
            ["Max Open Positions", "10", "Maximum number of open positions"],
            ["", "", ""],
            ["Instruments to Trade", "", ""],
            ["EUR_USD", "TRUE", "Enable EUR/USD trading"],
            ["GBP_USD", "TRUE", "Enable GBP/USD trading"],
            ["USD_JPY", "TRUE", "Enable USD/JPY trading"],
            ["USD_CHF", "TRUE", "Enable USD/CHF trading"],
            ["AUD_USD", "TRUE", "Enable AUD/USD trading"],
            ["USD_CAD", "TRUE", "Enable USD/CAD trading"],
            ["NZD_USD", "TRUE", "Enable NZD/USD trading"],
            ["EUR_GBP", "TRUE", "Enable EUR/GBP trading"],
        ]

        await worksheet.update("A1:C17", config_data)
        await worksheet.format(
            "A1:C1",
            {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.8, "green": 0.8, "blue": 0.2},
            },
        )

    async def _format_performance_sheet(self):
        """Formats the 'Performance Analytics' worksheet with headers."""
        spreadsheet = await self.get_spreadsheet()
        worksheet = await spreadsheet.worksheet(self.worksheet_names["performance"])

        headers = [
            ["Performance Analytics", "", "", ""],
            ["Period", "P&L", "Trades", "Win Rate"],
            ["Today", "", "", ""],
            ["This Week", "", "", ""],
            ["This Month", "", "", ""],
            ["", "", "", ""],
            ["Top Performing Pairs", "", "", ""],
            ["Symbol", "P&L", "Trades", "Win Rate"],
        ]

        await worksheet.update("A1:D8", headers)
        await worksheet.format(
            "A1:D1",
            {
                "textFormat": {"bold": True, "fontSize": 14},
                "backgroundColor": {"red": 0.2, "green": 0.8, "blue": 0.8},
            },
        )

    async def update_dashboard(self, account_summary: AccountSummary):
        """
        Updates the 'Dashboard' worksheet with the latest account summary.

        Args:
            account_summary (AccountSummary): The account summary data object.

        Raises:
            Exception: If the dashboard update fails.
        """
        try:
            spreadsheet = await self.get_spreadsheet()
            worksheet = await spreadsheet.worksheet(self.worksheet_names["dashboard"])

            # Update values
            updates = [
                ("B3", account_summary.balance),
                ("B4", account_summary.equity),
                ("B5", account_summary.free_margin),
                ("B6", f"{account_summary.margin_level:.2f}%"),
                ("B7", account_summary.total_open_positions),
                ("B8", account_summary.total_unrealized_pnl),
                ("B9", account_summary.daily_pnl),
                ("B10", account_summary.weekly_pnl),
                ("B11", account_summary.monthly_pnl),
                ("B14", f"{account_summary.max_drawdown:.2f}%"),
                ("B15", f"{account_summary.win_rate:.2f}%"),
                ("B16", account_summary.profit_factor),
                (
                    "B18",
                    account_summary.last_updated.strftime("%Y-%m-%d %H:%M:%S UTC"),
                ),
            ]

            # gspread batch_update is more efficient for multiple cell updates
            update_requests = [
                {"range": cell, "values": [[value]]} for cell, value in updates
            ]
            await worksheet.batch_update(update_requests)

            logger.info("Dashboard updated successfully")

        except Exception as e:
            logger.error(f"Error updating dashboard: {e}")
            raise

    async def update_positions(self, positions: List[Position]):
        """
        Updates the 'Open Positions' worksheet with the current list of open positions.

        This method clears the existing data and writes the new data.

        Args:
            positions (List[Position]): A list of current open positions.

        Raises:
            Exception: If updating positions fails.
        """
        try:
            spreadsheet = await self.get_spreadsheet()
            worksheet = await spreadsheet.worksheet(
                self.worksheet_names["open_positions"]
            )

            # Clear existing data (except headers)
            await worksheet.clear()

            # Re-add headers
            headers = [
                "Ticket",
                "Symbol",
                "Type",
                "Lots",
                "Open Price",
                "Current Price",
                "Stop Loss",
                "Take Profit",
                "Open Time",
                "Unrealized P&L",
                "Commission",
                "Swap",
            ]
            await worksheet.update("A1:L1", [headers])

            # Add position data
            if positions:
                rows = [
                    [
                        pos.ticket,
                        pos.symbol,
                        pos.type,
                        pos.lots,
                        pos.open_price,
                        pos.current_price,
                        pos.stop_loss or "",
                        pos.take_profit or "",
                        pos.open_time.strftime("%Y-%m-%d %H:%M:%S"),
                        pos.unrealized_pnl,
                        pos.commission,
                        pos.swap,
                    ]
                    for pos in positions
                ]
                await worksheet.update(f"A2:L{len(rows) + 1}", rows)

            logger.info(f"Updated {len(positions)} open positions")

        except Exception as e:
            logger.error(f"Error updating positions: {e}")
            raise

    async def add_closed_trade(self, trade: ClosedTrade):
        """
        Appends a new closed trade to the 'Closed Trades' worksheet.

        Args:
            trade (ClosedTrade): The closed trade data object.

        Raises:
            Exception: If adding the trade fails.
        """
        try:
            spreadsheet = await self.get_spreadsheet()
            worksheet = await spreadsheet.worksheet(
                self.worksheet_names["closed_trades"]
            )

            # Add trade data to the next empty row
            trade_data = [
                trade.ticket,
                trade.symbol,
                trade.type,
                trade.lots,
                trade.open_price,
                trade.close_price,
                trade.stop_loss or "",
                trade.take_profit or "",
                trade.open_time.strftime("%Y-%m-%d %H:%M:%S"),
                trade.close_time.strftime("%Y-%m-%d %H:%M:%S"),
                round(trade.duration_hours, 2),
                trade.realized_pnl,
                trade.commission,
                trade.swap,
            ]

            await worksheet.append_row(trade_data)
            logger.info(f"Added closed trade: {trade.ticket}")

        except Exception as e:
            logger.error(f"Error adding closed trade: {e}")
            raise

    async def get_risk_parameters(self) -> RiskParameters:
        """
        Retrieves risk parameters from the 'Risk Configuration' worksheet.

        Returns:
            RiskParameters: A data object containing the configured risk parameters.
                            Returns default parameters if retrieval fails.
        """
        try:
            spreadsheet = await self.get_spreadsheet()
            worksheet = await spreadsheet.worksheet(self.worksheet_names["risk_config"])

            values = await worksheet.get_all_values()

            # Extract risk parameters from the sheet
            params_dict = {}
            instruments = []
            is_instrument_section = False

            for row in values[1:]:  # Skip header
                if not row:
                    continue
                param_name = row[0].strip()

                if "Instruments to Trade" in param_name:
                    is_instrument_section = True
                    continue

                if is_instrument_section:
                    if len(row) > 1 and row[1].strip().upper() == "TRUE":
                        instruments.append(param_name)
                elif len(row) > 1:
                    params_dict[param_name] = row[1].strip()

            risk_params = RiskParameters(
                max_risk_per_trade=float(params_dict.get("Max Risk Per Trade", 0.02)),
                max_daily_drawdown=float(params_dict.get("Max Daily Drawdown", 0.05)),
                max_correlation=float(params_dict.get("Max Correlation", 0.7)),
                max_exposure_per_currency=float(
                    params_dict.get("Max Exposure Per Currency", 0.1)
                ),
                max_lot_size=float(params_dict.get("Max Lot Size", 1.0)),
                max_open_positions=int(params_dict.get("Max Open Positions", 10)),
                instruments_to_trade=instruments
                or None,  # Use None to trigger defaults if empty
            )

            return risk_params

        except Exception as e:
            logger.error(f"Error getting risk parameters: {e}")
            return RiskParameters()  # Return defaults on failure


class ExcelManager:
    """
    Manages a local Excel file for portfolio tracking as an alternative to Google Sheets.

    Attributes:
        file_path (str): The path to the Excel workbook file.
    """

    def __init__(self, file_path: str):
        """
        Initializes the ExcelManager and ensures the file exists.

        Args:
            file_path (str): The path to the Excel file.
        """
        self.file_path = file_path
        self.ensure_file_exists()

    def ensure_file_exists(self):
        """Creates the Excel file with the required worksheets if it doesn't exist."""
        if not os.path.exists(self.file_path):
            workbook = openpyxl.Workbook()

            # Create worksheets
            workbook.create_sheet("Dashboard")
            workbook.create_sheet("Open Positions")
            workbook.create_sheet("Closed Trades")
            workbook.create_sheet("Risk Configuration")
            workbook.create_sheet("Performance Analytics")

            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            workbook.save(self.file_path)
            logger.info(f"Created new Excel file: {self.file_path}")

    def update_dashboard(self, account_summary: AccountSummary):
        """
        Updates the 'Dashboard' worksheet in the Excel file.

        Args:
            account_summary (AccountSummary): The account summary data object.

        Raises:
            Exception: If updating the Excel dashboard fails.
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook["Dashboard"]

            # Clear existing content
            worksheet.delete_rows(1, worksheet.max_row)

            # Add headers and data
            data = [
                ["Account Summary", "", "", ""],
                ["Metric", "Value", "Previous", "Change"],
                ["Balance", account_summary.balance, "", ""],
                ["Equity", account_summary.equity, "", ""],
                ["Free Margin", account_summary.free_margin, "", ""],
                [
                    "Margin Level",
                    (
                        f"{account_summary.margin_level:.2f}%"
                        if account_summary.margin_level
                        else "N/A"
                    ),
                    "",
                    "",
                ],
                ["Open Positions", account_summary.total_open_positions, "", ""],
                ["Unrealized P&L", account_summary.total_unrealized_pnl, "", ""],
                ["Daily P&L", account_summary.daily_pnl, "", ""],
                ["Weekly P&L", account_summary.weekly_pnl, "", ""],
                ["Monthly P&L", account_summary.monthly_pnl, "", ""],
                ["", "", "", ""],
                ["Risk Metrics", "", "", ""],
                ["Max Drawdown", f"{account_summary.max_drawdown:.2f}%", "", ""],
                ["Win Rate", f"{account_summary.win_rate:.2f}%", "", ""],
                ["Profit Factor", account_summary.profit_factor, "", ""],
                ["", "", "", ""],
                [
                    "Last Updated",
                    account_summary.last_updated.strftime("%Y-%m-%d %H:%M:%S UTC"),
                    "",
                    "",
                ],
            ]

            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value

                    # Format headers
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14)
                        cell.fill = PatternFill(
                            start_color="3366CC", end_color="3366CC", fill_type="solid"
                        )

            workbook.save(self.file_path)
            logger.info("Excel dashboard updated successfully")

        except Exception as e:
            logger.error(f"Error updating Excel dashboard: {e}")
            raise

    def update_positions(self, positions: List[Position]):
        """
        Updates the 'Open Positions' worksheet in the Excel file.

        Args:
            positions (List[Position]): A list of current open positions.

        Raises:
            Exception: If updating Excel positions fails.
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook["Open Positions"]

            # Clear existing content
            worksheet.delete_rows(1, worksheet.max_row)

            # Add headers
            headers = [
                "Ticket",
                "Symbol",
                "Type",
                "Lots",
                "Open Price",
                "Current Price",
                "Stop Loss",
                "Take Profit",
                "Open Time",
                "Unrealized P&L",
                "Commission",
                "Swap",
            ]

            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CC3333", end_color="CC3333", fill_type="solid"
                )

            # Add position data
            for row_idx, pos in enumerate(positions, 2):
                row_data = [
                    pos.ticket,
                    pos.symbol,
                    pos.type,
                    pos.lots,
                    pos.open_price,
                    pos.current_price,
                    pos.stop_loss or "",
                    pos.take_profit or "",
                    pos.open_time.strftime("%Y-%m-%d %H:%M:%S"),
                    pos.unrealized_pnl,
                    pos.commission,
                    pos.swap,
                ]

                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx).value = value

            workbook.save(self.file_path)
            logger.info(f"Updated {len(positions)} positions in Excel")

        except Exception as e:
            logger.error(f"Error updating Excel positions: {e}")
            raise


class AssetManager:
    """
    The main service for managing trading assets, portfolio, and risk.

    This class acts as a facade, orchestrating either the GoogleSheetsManager or
    the ExcelManager based on the configuration.

    Attributes:
        use_google_sheets (bool): Flag to determine whether to use Google Sheets or Excel.
        sheets_manager (Optional[GoogleSheetsManager]): Instance of the Google Sheets manager.
        excel_manager (Optional[ExcelManager]): Instance of the Excel manager.
        risk_parameters (RiskParameters): The currently loaded risk parameters.
    """

    def __init__(self, use_google_sheets: bool = True):
        """
        Initializes the AssetManager.

        Args:
            use_google_sheets (bool): If True, Google Sheets will be used.
                                      Otherwise, a local Excel file is used.
        """
        self.use_google_sheets = use_google_sheets
        self.sheets_manager = None
        self.excel_manager = None
        self.risk_parameters = RiskParameters()

        if self.use_google_sheets:
            if (
                settings.GOOGLE_SHEETS_CREDENTIALS_PATH
                and settings.GOOGLE_SHEETS_SPREADSHEET_KEY
            ):
                self.sheets_manager = GoogleSheetsManager(
                    credentials_path=settings.GOOGLE_SHEETS_CREDENTIALS_PATH,
                    spreadsheet_key=settings.GOOGLE_SHEETS_SPREADSHEET_KEY,
                )
            else:
                logger.warning(
                    "Google Sheets is enabled but credentials or key are missing."
                )
                self.use_google_sheets = False  # Fallback to Excel

        if not self.use_google_sheets:
            self.excel_manager = ExcelManager(
                file_path=settings.EXCEL_FILE_PATH or "trading_portfolio.xlsx"
            )

    async def initialize(self):
        """
        Initializes the selected asset management backend (Sheets or Excel).

        For Google Sheets, it also fetches the initial risk parameters.

        Raises:
            Exception: If initialization fails.
        """
        try:
            if self.use_google_sheets and self.sheets_manager:
                await self.sheets_manager.initialize()
                await self.sheets_manager.setup_worksheets()
                self.risk_parameters = await self.sheets_manager.get_risk_parameters()

            logger.info("Asset manager initialized successfully")

        except Exception as e:
            logger.error(f"Failed to initialize asset manager: {e}")
            raise

    async def update_portfolio(
        self, account_summary: AccountSummary, positions: List[Position]
    ):
        """
        Updates the portfolio dashboard and open positions list.

        Args:
            account_summary (AccountSummary): The latest account summary.
            positions (List[Position]): The current list of open positions.

        Raises:
            Exception: If the portfolio update fails.
        """
        try:
            if self.use_google_sheets and self.sheets_manager:
                await self.sheets_manager.update_dashboard(account_summary)
                await self.sheets_manager.update_positions(positions)
            elif self.excel_manager:
                self.excel_manager.update_dashboard(account_summary)
                self.excel_manager.update_positions(positions)

            logger.info("Portfolio updated successfully")

        except Exception as e:
            logger.error(f"Error updating portfolio: {e}")
            raise

    async def log_trade(self, trade: ClosedTrade):
        """
        Logs a closed trade to the appropriate backend.

        Args:
            trade (ClosedTrade): The trade to be logged.

        Raises:
            Exception: If logging the trade fails.
        """
        try:
            if self.use_google_sheets and self.sheets_manager:
                await self.sheets_manager.add_closed_trade(trade)
            # TODO: Implement trade logging for ExcelManager
            # elif self.excel_manager:
            #     self.excel_manager.add_closed_trade(trade)

            logger.info(f"Trade logged: {trade.ticket}")

        except Exception as e:
            logger.error(f"Error logging trade: {e}")
            raise

    def validate_trade_risk(
        self, signal, current_equity: float, open_positions: List[Position]
    ) -> bool:
        """
        Validates if a potential trade conforms to the loaded risk parameters.

        Args:
            signal: A signal object containing trade details like instrument and volume.
            current_equity (float): The current equity of the account.
            open_positions (List[Position]): The list of current open positions.

        Returns:
            bool: True if the trade is within risk parameters, False otherwise.
        """
        try:
            # Check if instrument is allowed
            if signal.instrument not in self.risk_parameters.instruments_to_trade:
                logger.warning(f"Instrument {signal.instrument} not in allowed list")
                return False

            # Check maximum open positions
            if len(open_positions) >= self.risk_parameters.max_open_positions:
                logger.warning("Maximum open positions reached")
                return False

            # Check maximum lot size
            if signal.volume > self.risk_parameters.max_lot_size:
                logger.warning(f"Volume {signal.volume} exceeds maximum lot size")
                return False

            # Check risk per trade
            if signal.stop_loss and current_equity > 0:
                # This is an approximation. Accurate calculation requires pip value.
                risk_amount = abs(
                    signal.volume * (signal.open_price - signal.stop_loss) * 100000
                )
                risk_percentage = risk_amount / current_equity

                if risk_percentage > self.risk_parameters.max_risk_per_trade:
                    logger.warning(
                        f"Risk {risk_percentage:.2%} exceeds maximum risk per trade"
                    )
                    return False

            return True

        except Exception as e:
            logger.error(f"Error validating trade risk: {e}")
            return False

    def get_risk_parameters(self) -> RiskParameters:
        """
        Returns the currently loaded risk parameters.

        Returns:
            RiskParameters: The current risk parameters.
        """
        return self.risk_parameters


# Factory function
async def create_asset_manager(use_google_sheets: bool = True) -> AssetManager:
    """
    Factory function to create and initialize an instance of the AssetManager.

    Args:
        use_google_sheets (bool): Specifies whether to use Google Sheets or Excel.

    Returns:
        AssetManager: An initialized instance of the AssetManager.
    """
    manager = AssetManager(use_google_sheets=use_google_sheets)
    await manager.initialize()
    return manager

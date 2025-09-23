#!/usr/bin/env python3
"""
ForexConnect Excel Integration for GenX FX Trading System
Connects to real FXCM data via ForexConnect API using demo credentials
Generates Excel dashboard with real market data and signals
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import sys
import time

# Add project root to path
sys.path.append(str(os.path.dirname(__file__)))

try:
    import forexconnect as fx
    FOREXCONNECT_AVAILABLE = True
    print("✅ ForexConnect module found")
except ImportError:
    FOREXCONNECT_AVAILABLE = False
    print("⚠️ ForexConnect module not found - will use mock data")

class ForexConnectExcelGenerator:
    """
    Connects to the ForexConnect API to fetch live market data, generate trading signals,
    and create an Excel dashboard. Falls back to mock data if the API is unavailable.
    """
    def __init__(self):
        """
        Initializes the generator, setting up API credentials, paths, and pairs.
        """
        self.pairs = ['EUR/USD', 'GBP/USD', 'USD/JPY', 'USD/CHF', 'AUD/USD', 'USD/CAD', 'NZD/USD']
        self.signal_output_dir = 'signal_output'
        self.session = None
        self.connected = False
        
        # FXCM ForexConnect configuration
        self.url = os.getenv('FXCM_URL', 'https://api-demo.fxcm.com')
        self.username = os.getenv('FXCM_USERNAME', 'demo')
        self.password = os.getenv('FXCM_PASSWORD')
        if not self.password:
            raise ValueError("FXCM_PASSWORD environment variable is required")
        self.connection_type = os.getenv('FXCM_CONNECTION_TYPE', 'Demo')
        
        # Create output directory
        os.makedirs(self.signal_output_dir, exist_ok=True)
        
    def connect_forexconnect(self) -> bool:
        """
        Connects to the ForexConnect API using credentials from environment variables.

        Returns:
            bool: True if the connection is successful, False otherwise.
        """
        if not FOREXCONNECT_AVAILABLE:
            print("⚠️ ForexConnect not available - using mock data")
            return False
            
        try:
            print(f"🔌 Connecting to FXCM ForexConnect...")
            print(f"   URL: {self.url}")
            print(f"   Connection: {self.connection_type}")
            print(f"   Username: {self.username}")
            
            # Create session
            self.session = fx.O2GSession()
            
            # Create session descriptor
            session_descriptor = fx.O2GSessionDescriptor()
            session_descriptor.setUrl(self.url)
            session_descriptor.setUser(self.username)
            session_descriptor.setPassword(self.password)
            session_descriptor.setConnection(self.connection_type)
            
            # Login with timeout
            print("🔑 Logging in...")
            status = self.session.login(session_descriptor.getUser(),
                                      session_descriptor.getPassword(),
                                      session_descriptor.getUrl(),
                                      session_descriptor.getConnection())
            
            if status == fx.Connected:
                self.connected = True
                print("✅ Connected to FXCM successfully!")
                return True
            else:
                print(f"❌ Connection failed with status: {status}")
                return False
                
        except Exception as e:
            print(f"❌ ForexConnect connection error: {e}")
            return False
    
    def get_live_prices(self) -> dict:
        """
        Gets live market prices from ForexConnect. Falls back to mock data if not connected.

        Returns:
            dict: A dictionary of prices for each symbol.
        """
        if not self.connected:
            return self.get_mock_prices()
            
        try:
            print("📊 Fetching live market data...")
            
            # Get market data table
            table_manager = self.session.getTableManager()
            offers_table = table_manager.getTable(fx.Offers)
            
            prices = {}
            for i in range(offers_table.size()):
                offer_row = offers_table.getRow(i)
                symbol = offer_row.getInstrument()
                
                if symbol in self.pairs:
                    bid = offer_row.getBid()
                    ask = offer_row.getAsk()
                    spread = ask - bid
                    
                    prices[symbol] = {
                        'bid': bid,
                        'ask': ask,
                        'spread': spread,
                        'time': datetime.now()
                    }
                    
            print(f"✅ Retrieved prices for {len(prices)} pairs")
            return prices
            
        except Exception as e:
            print(f"⚠️ Error getting live prices: {e}")
            return self.get_mock_prices()
    
    def get_mock_prices(self) -> dict:
        """
        Generates fallback mock prices when ForexConnect is unavailable.

        Returns:
            dict: A dictionary of mock prices for each symbol.
        """
        print("📊 Using mock price data...")
        
        base_prices = {
            'EUR/USD': 1.10500, 'GBP/USD': 1.27000, 'USD/JPY': 149.50,
            'USD/CHF': 0.88200, 'AUD/USD': 0.65800, 'USD/CAD': 1.36500, 'NZD/USD': 0.58900
        }
        
        prices = {}
        for symbol, base_price in base_prices.items():
            # Add realistic spread
            spread = 0.0001 if 'JPY' not in symbol else 0.01
            spread *= np.random.uniform(1.5, 3.0)  # Realistic spread variation
            
            # Add small price movement
            variation = np.random.uniform(-0.002, 0.002)
            current_price = base_price * (1 + variation)
            
            bid = current_price - spread/2
            ask = current_price + spread/2
            
            prices[symbol] = {
                'bid': round(bid, 5),
                'ask': round(ask, 5),
                'spread': round(spread, 5),
                'time': datetime.now()
            }
        
        return prices
    
    def generate_signals_from_prices(self, prices: dict, num_signals: int = 10) -> pd.DataFrame:
        """
        Generates trading signals based on live or mock market data.

        Args:
            prices (dict): A dictionary of market prices.
            num_signals (int, optional): The number of signals to generate. Defaults to 10.

        Returns:
            pd.DataFrame: A DataFrame containing the generated signals.
        """
        signals = []
        
        for i in range(num_signals):
            # Select random symbol
            symbol = np.random.choice(list(prices.keys()))
            symbol_clean = symbol.replace('/', '')  # Convert EUR/USD to EURUSD
            
            price_data = prices[symbol]
            current_price = (price_data['bid'] + price_data['ask']) / 2
            spread = price_data['spread']
            
            # Generate signal based on simple technical logic
            signal_type = self.determine_signal_type(current_price, symbol)
            
            # Calculate entry price with spread consideration
            if signal_type == 'BUY':
                entry_price = price_data['ask']  # Buy at ask
                stop_loss = entry_price - self.calculate_atr_distance(symbol, current_price)
                take_profit = entry_price + self.calculate_atr_distance(symbol, current_price) * 2
            else:
                entry_price = price_data['bid']  # Sell at bid
                stop_loss = entry_price + self.calculate_atr_distance(symbol, current_price)
                take_profit = entry_price - self.calculate_atr_distance(symbol, current_price) * 2
            
            # Calculate realistic confidence based on spread and volatility
            confidence = self.calculate_confidence(symbol, spread, current_price)
            risk_reward = abs(take_profit - entry_price) / abs(entry_price - stop_loss)
            
            signal = {
                'Timestamp': datetime.now() + timedelta(minutes=np.random.randint(-5, 5)),
                'Symbol': symbol_clean,
                'Signal': signal_type,
                'Entry_Price': round(entry_price, 5),
                'Stop_Loss': round(stop_loss, 5),
                'Take_Profit': round(take_profit, 5),
                'Bid': price_data['bid'],
                'Ask': price_data['ask'],
                'Spread': spread,
                'Lot_Size': round(np.random.uniform(0.01, 0.10), 2),
                'Confidence': confidence,
                'Risk_Reward': round(risk_reward, 2),
                'Magic_Number': 123450 + i,
                'Timeframe': np.random.choice(['M15', 'H1', 'H4']),
                'Status': np.random.choice(['Active', 'Pending'], p=[0.7, 0.3]),
                'Comment': f'GenX_{signal_type}_{symbol_clean}_Live',
                'Data_Source': 'ForexConnect' if self.connected else 'Mock'
            }
            
            signals.append(signal)
        
        return pd.DataFrame(signals)
    
    def determine_signal_type(self, price: float, symbol: str) -> str:
        """
        Determines the signal type (BUY/SELL) based on a simple logic.
        This is a placeholder and should be replaced with an actual trading strategy.

        Args:
            price (float): The current price of the symbol.
            symbol (str): The financial instrument symbol.

        Returns:
            str: The signal type, either 'BUY' or 'SELL'.
        """
        # Simple random with slight bias based on time
        hour = datetime.now().hour
        
        # London session bias (more BUY signals)
        if 8 <= hour <= 12:
            return np.random.choice(['BUY', 'SELL'], p=[0.6, 0.4])
        # New York session bias (more SELL signals) 
        elif 13 <= hour <= 17:
            return np.random.choice(['BUY', 'SELL'], p=[0.4, 0.6])
        else:
            return np.random.choice(['BUY', 'SELL'])
    
    def calculate_atr_distance(self, symbol: str, price: float) -> float:
        """
        Calculates a simplified ATR-based distance for stop loss.
        This is a placeholder for a real ATR calculation.

        Args:
            symbol (str): The financial instrument symbol.
            price (float): The current price of the symbol.

        Returns:
            float: The calculated distance for the stop loss.
        """
        # Simplified ATR calculation - replace with actual ATR
        base_atr = {
            'EUR/USD': 0.0015, 'GBP/USD': 0.0020, 'USD/JPY': 0.15,
            'USD/CHF': 0.0012, 'AUD/USD': 0.0018, 'USD/CAD': 0.0016, 'NZD/USD': 0.0020
        }
        
        symbol_key = symbol.replace('/', '')
        if 'JPY' in symbol_key:
            return base_atr.get(symbol, 0.15) * np.random.uniform(0.8, 1.2)
        else:
            return base_atr.get(symbol, 0.0015) * np.random.uniform(0.8, 1.2)
    
    def calculate_confidence(self, symbol: str, spread: float, price: float) -> float:
        """
        Calculates signal confidence based on market conditions like spread and time.

        Args:
            symbol (str): The financial instrument symbol.
            spread (float): The current spread.
            price (float): The current price.

        Returns:
            float: The calculated confidence score between 0.55 and 0.95.
        """
        # Higher confidence for tighter spreads
        spread_score = max(0.5, 1.0 - (spread * 10000))  # Convert to pips
        
        # Time-based confidence (higher during major sessions)
        hour = datetime.now().hour
        time_score = 0.8
        if 8 <= hour <= 17:  # London + NY overlap
            time_score = 0.9
        elif hour < 6 or hour > 20:  # Asian session
            time_score = 0.6
            
        # Combine factors
        base_confidence = spread_score * time_score
        final_confidence = np.random.uniform(base_confidence * 0.8, base_confidence * 1.0)
        
        return round(np.clip(final_confidence, 0.55, 0.95), 2)
    
    def create_excel_dashboard(self, df: pd.DataFrame, prices: dict) -> str:
        """
        Creates an Excel dashboard with live market data and generated signals.

        Args:
            df (pd.DataFrame): DataFrame containing the generated signals.
            prices (dict): Dictionary of live market prices.

        Returns:
            str: The file path of the created Excel dashboard.
        """
        file_path = os.path.join(self.signal_output_dir, 'genx_live_signals.xlsx')
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create worksheets
        ws_active = wb.create_sheet('Live Signals')
        ws_prices = wb.create_sheet('Market Prices')
        ws_summary = wb.create_sheet('Dashboard')
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        buy_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        sell_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        live_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # === LIVE SIGNALS SHEET ===
        headers = ['Time', 'Symbol', 'Signal', 'Entry', 'Stop Loss', 'Take Profit', 
                  'Bid', 'Ask', 'Spread', 'Confidence', 'R:R', 'Source']
        
        for col, header in enumerate(headers, 1):
            cell = ws_active.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, signal) in enumerate(df.iterrows(), 2):
            ws_active.cell(row=row_idx, column=1, value=signal['Timestamp'].strftime('%H:%M:%S'))
            ws_active.cell(row=row_idx, column=2, value=signal['Symbol'])
            ws_active.cell(row=row_idx, column=3, value=signal['Signal'])
            ws_active.cell(row=row_idx, column=4, value=signal['Entry_Price'])
            ws_active.cell(row=row_idx, column=5, value=signal['Stop_Loss'])
            ws_active.cell(row=row_idx, column=6, value=signal['Take_Profit'])
            ws_active.cell(row=row_idx, column=7, value=signal['Bid'])
            ws_active.cell(row=row_idx, column=8, value=signal['Ask'])
            ws_active.cell(row=row_idx, column=9, value=signal['Spread'])
            ws_active.cell(row=row_idx, column=10, value=f"{signal['Confidence']:.1%}")
            ws_active.cell(row=row_idx, column=11, value=signal['Risk_Reward'])
            ws_active.cell(row=row_idx, column=12, value=signal['Data_Source'])
            
            # Color coding
            signal_cell = ws_active.cell(row=row_idx, column=3)
            if signal['Signal'] == 'BUY':
                signal_cell.fill = buy_fill
            else:
                signal_cell.fill = sell_fill
                
            # Highlight live data
            if signal['Data_Source'] == 'ForexConnect':
                ws_active.cell(row=row_idx, column=12).fill = live_fill
        
        # Auto-adjust columns
        for col in range(1, 13):
            ws_active.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # === MARKET PRICES SHEET ===
        price_headers = ['Symbol', 'Bid', 'Ask', 'Spread (pips)', 'Last Update']
        for col, header in enumerate(price_headers, 1):
            cell = ws_prices.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (symbol, price_data) in enumerate(prices.items(), 2):
            spread_pips = price_data['spread'] * (10000 if 'JPY' not in symbol else 100)
            
            ws_prices.cell(row=row_idx, column=1, value=symbol)
            ws_prices.cell(row=row_idx, column=2, value=price_data['bid'])
            ws_prices.cell(row=row_idx, column=3, value=price_data['ask'])
            ws_prices.cell(row=row_idx, column=4, value=round(spread_pips, 1))
            ws_prices.cell(row=row_idx, column=5, value=price_data['time'].strftime('%H:%M:%S'))
        
        # === DASHBOARD SHEET ===
        ws_summary.cell(row=1, column=1, value='GenX Live Trading Dashboard')
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        connection_status = "🟢 LIVE (ForexConnect)" if self.connected else "🟡 DEMO (Mock Data)"
        
        summary_data = [
            ['Connection Status', connection_status],
            ['Data Source', 'ForexConnect API' if self.connected else 'Mock Data'],
            ['Total Signals', len(df)],
            ['Live Signals', len(df[df['Data_Source'] == 'ForexConnect'])],
            ['Average Confidence', f"{df['Confidence'].mean():.1%}"],
            ['Average Spread', f"{df['Spread'].mean():.1f} pips"],
            ['Last Update', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        wb.save(file_path)
        print(f"✅ Live Excel dashboard: {file_path}")
        return file_path
    
    def disconnect(self):
        """
        Disconnects from the ForexConnect session if it's active.
        """
        if self.session and self.connected:
            try:
                self.session.logout()
                print("✅ Disconnected from ForexConnect")
            except Exception as e:
                print(f"⚠️ Disconnect error: {e}")
    
    def run_live_demo(self, num_signals: int = 10) -> tuple:
        """
        Runs the live demo, connecting to ForexConnect, generating signals,
        and creating all output files.

        Args:
            num_signals (int, optional): The number of signals to generate. Defaults to 10.

        Returns:
            tuple: A tuple containing the file paths of the created output files.
        """
        print("🚀 GenX FX - Live ForexConnect Demo")
        print("=" * 50)
        
        try:
            # Connect to ForexConnect
            if FOREXCONNECT_AVAILABLE:
                self.connect_forexconnect()
            
            # Get live prices
            prices = self.get_live_prices()
            
            # Generate signals from live data
            print(f"📊 Generating {num_signals} signals from live market data...")
            df = self.generate_signals_from_prices(prices, num_signals)
            
            # Create Excel dashboard
            excel_file = self.create_excel_dashboard(df, prices)
            
            # Create standard outputs
            from demo_excel_generator import ForexSignalGenerator
            generator = ForexSignalGenerator()
            mt4_file = generator.create_mt4_csv(df)
            mt5_file = generator.create_mt5_csv(df)
            json_file = generator.create_json_output(df)
            
            # Print summary
            print(f"\n📈 Live Trading Summary:")
            print(f"   • Data Source: {'ForexConnect (LIVE)' if self.connected else 'Mock Data'}")
            print(f"   • Total Signals: {len(df)}")
            print(f"   • Average Spread: {df['Spread'].mean():.1f} pips")
            print(f"   • Average Confidence: {df['Confidence'].mean():.1%}")
            
            print(f"\n📁 Output Files:")
            print(f"   📊 Live Dashboard: {excel_file}")
            print(f"   📈 MT4 Signals: {mt4_file}")
            print(f"   📈 MT5 Signals: {mt5_file}")
            print(f"   🔗 JSON API: {json_file}")
            
            return excel_file, mt4_file, mt5_file, json_file
            
        finally:
            self.disconnect()

if __name__ == "__main__":
    generator = ForexConnectExcelGenerator()
    generator.run_live_demo()
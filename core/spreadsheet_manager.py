"""
Spreadsheet Manager for MT4/5 Signal Output
Exports trading signals to Excel/CSV files for EA consumption
"""

import asyncio
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import os
from pathlib import Path
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import csv

logger = logging.getLogger(__name__)


class SpreadsheetManager:
    """
    Manages the output of trading signals to various spreadsheet formats.

    This class is designed to export trading signals to formats easily consumable
    by MetaTrader 4/5 Expert Advisors (EAs), such as Excel, CSV, and JSON.

    Attributes:
        config (Dict[str, Any]): Configuration settings.
        output_dir (Path): The directory for output files.
        active_signals (Dict): A dictionary of currently active signals.
        signal_history (List): A history of all processed signals.
        excel_file (Path): The path to the main Excel output file.
        csv_file (Path): The path to the main CSV output file.
        json_file (Path): The path to the JSON output file.
        mt4_file (Path): The path to the MT4-specific CSV file.
        mt5_file (Path): The path to the MT5-specific CSV file.
    """

    def __init__(self, config: Dict[str, Any]):
        """
        Initializes the SpreadsheetManager.

        Args:
            config (Dict[str, Any]): A dictionary of configuration parameters,
                                     including 'output_directory'.
        """
        self.config = config
        self.output_dir = Path(config.get("output_directory", "signal_output"))
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Signal storage
        self.active_signals: Dict[str, Any] = {}
        self.signal_history: List[Dict[str, Any]] = []
        self.performance_data: Dict[str, Any] = {}

        # File paths
        self.excel_file = self.output_dir / "genx_signals.xlsx"
        self.csv_file = self.output_dir / "genx_signals.csv"
        self.json_file = self.output_dir / "genx_signals.json"
        self.mt4_file = self.output_dir / "MT4_Signals.csv"
        self.mt5_file = self.output_dir / "MT5_Signals.csv"

        # Backup directory
        self.backup_dir = self.output_dir / "backups"
        self.backup_dir.mkdir(exist_ok=True)

        self.last_update: Optional[datetime] = None
        self.update_interval = config.get("update_interval", 30)  # seconds
        self.max_signals = config.get("max_signals", 50)

        logger.info(f"Spreadsheet Manager initialized - Output: {self.output_dir}")

    async def initialize(self):
        """
        Initializes the spreadsheet manager.

        This involves creating the initial spreadsheet files if they don't exist
        and loading any previously saved signals.

        Raises:
            Exception: If initialization fails.
        """
        try:
            await self._create_initial_files()
            await self._load_existing_signals()
            logger.info("Spreadsheet Manager initialization complete")
        except Exception as e:
            logger.error(f"Error initializing spreadsheet manager: {e}")
            raise

    async def update_signals(self, signals: List[Any]):
        """
        Updates and exports a list of new signals to all configured formats.

        Args:
            signals (List[Any]): A list of signal objects to be processed.
        """
        try:
            current_time = datetime.now()

            for signal in signals:
                signal_data = (
                    signal.to_mt4_format()
                    if hasattr(signal, "to_mt4_format")
                    else signal
                )
                signal_id = signal_data.get(
                    "Magic", str(hash(f"{signal_data.get('Symbol')}_{current_time}"))
                )

                self.active_signals[signal_id] = {
                    **signal_data,
                    "ID": signal_id,
                    "Status": "ACTIVE",
                    "CreatedTime": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "LastUpdate": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                }
                self.signal_history.append(self.active_signals[signal_id].copy())

            # Clean up old signals
            await self._cleanup_old_signals()

            # Update all output files
            await self._update_excel_file()
            await self._update_csv_files()
            await self._update_json_file()
            await self._update_mt4_mt5_files()

            # Create backup if needed
            if self.config.get("backup_enabled", True):
                await self._create_backup_if_needed()

            self.last_update = current_time
            logger.info(f"Updated {len(signals)} signals across all formats")

        except Exception as e:
            logger.error(f"Error updating signals: {e}")

    async def _create_initial_files(self):
        """Creates the initial signal files with headers if they don't exist."""
        if not self.excel_file.exists():
            await self._create_excel_template()

        # CSV files
        csv_headers = [
            "ID",
            "Symbol",
            "Signal",
            "Strength",
            "EntryPrice",
            "StopLoss",
            "TakeProfit",
            "Confidence",
            "RiskReward",
            "PositionSize",
            "MaxRisk",
            "Timeframe",
            "Timestamp",
            "ExpiryTime",
            "MarketCondition",
            "TechnicalConfluence",
            "FundamentalScore",
            "Status",
            "CreatedTime",
            "LastUpdate",
        ]

        for file_path in [self.csv_file, self.mt4_file, self.mt5_file]:
            if not file_path.exists():
                with open(file_path, "w", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(csv_headers)

        # JSON file
        if not self.json_file.exists():
            with open(self.json_file, "w") as f:
                json.dump(
                    {"signals": [], "last_update": None, "metadata": {}}, f, indent=2
                )

    async def _create_excel_template(self):
        """Creates a new Excel file with pre-formatted sheets and headers."""
        try:
            wb = openpyxl.Workbook()

            # Main signals sheet
            ws_signals = wb.active
            if ws_signals:
                ws_signals.title = "Active Signals"

            # Headers
            headers = [
                "ID",
                "Symbol",
                "Signal",
                "Strength",
                "Entry Price",
                "Stop Loss",
                "Take Profit",
                "Confidence",
                "Risk/Reward",
                "Position Size %",
                "Max Risk %",
                "Timeframe",
                "Timestamp",
                "Expiry Time",
                "Market Condition",
                "Technical Confluence",
                "Fundamental Score",
                "Status",
                "Created Time",
                "Last Update",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws_signals.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Set column widths
            column_widths = [
                8,
                10,
                8,
                10,
                12,
                12,
                12,
                12,
                12,
                12,
                10,
                10,
                18,
                18,
                15,
                12,
                12,
                10,
                18,
                18,
            ]
            for col, width in enumerate(column_widths, 1):
                ws_signals.column_dimensions[
                    openpyxl.utils.get_column_letter(col)
                ].width = width

            # Signal History sheet
            ws_history = wb.create_sheet("Signal History")
            for col, header in enumerate(headers, 1):
                cell = ws_history.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Performance sheet
            ws_performance = wb.create_sheet("Performance")
            perf_headers = [
                "Symbol",
                "Total Signals",
                "Win Rate %",
                "Avg Confidence",
                "Avg Risk/Reward",
                "Last Signal",
            ]
            for col, header in enumerate(perf_headers, 1):
                cell = ws_performance.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="GenX FX Trading Signals").font = (
                Font(size=16, bold=True)
            )
            ws_summary.cell(
                row=2,
                column=1,
                value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            )

            summary_data = [
                ["Total Active Signals:", "=COUNTA('Active Signals'!A:A)-1"],
                ["Last Update:", ""],
                ["System Status:", "ACTIVE"],
                [""],
                ["Signal Strength Distribution:"],
                ["VERY_STRONG:", "=COUNTIF('Active Signals'!D:D,\"4\")"],
                ["STRONG:", "=COUNTIF('Active Signals'!D:D,\"3\")"],
                ["MODERATE:", "=COUNTIF('Active Signals'!D:D,\"2\")"],
                ["WEAK:", "=COUNTIF('Active Signals'!D:D,\"1\")"],
            ]

            for row, (label, formula) in enumerate(summary_data, 4):
                ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
                if formula:
                    ws_summary.cell(row=row, column=2, value=formula)

            wb.save(self.excel_file)
            logger.info("Created Excel template")

        except Exception as e:
            logger.error(f"Error creating Excel template: {e}")

    async def _update_excel_file(self):
        """Updates the Excel file with the current set of active signals and performance."""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws_signals = wb["Active Signals"]

            # Clear existing data (keep headers)
            ws_signals.delete_rows(2, ws_signals.max_row)

            # Add active signals
            for row, signal in enumerate(self.active_signals.values(), 2):
                data = [
                    signal.get("ID", ""),
                    signal.get("Symbol", ""),
                    signal.get("Signal", ""),
                    signal.get("Strength", ""),
                    signal.get("EntryPrice", 0),
                    signal.get("StopLoss", 0),
                    signal.get("TakeProfit", 0),
                    signal.get("Confidence", 0),
                    signal.get("RiskReward", 0),
                    signal.get("PositionSize", 0),
                    signal.get("MaxRisk", 0),
                    signal.get("Timeframe", ""),
                    signal.get("Timestamp", ""),
                    signal.get("ExpiryTime", ""),
                    signal.get("MarketCondition", ""),
                    signal.get("TechnicalConfluence", 0),
                    signal.get("FundamentalScore", 0),
                    signal.get("Status", ""),
                    signal.get("CreatedTime", ""),
                    signal.get("LastUpdate", ""),
                ]

                for col, value in enumerate(data, 1):
                    cell = ws_signals.cell(row=row, column=col, value=value)

                    # Color code by signal type
                    if col == 3:  # Signal column
                        if value == "BUY":
                            cell.fill = PatternFill(
                                start_color="C6EFCE",
                                end_color="C6EFCE",
                                fill_type="solid",
                            )
                        elif value == "SELL":
                            cell.fill = PatternFill(
                                start_color="FFC7CE",
                                end_color="FFC7CE",
                                fill_type="solid",
                            )

                    # Color code by strength
                    elif col == 4:  # Strength column
                        if value == 4:
                            cell.fill = PatternFill(
                                start_color="00B050",
                                end_color="00B050",
                                fill_type="solid",
                            )
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif value == 3:
                            cell.fill = PatternFill(
                                start_color="92D050",
                                end_color="92D050",
                                fill_type="solid",
                            )
                        elif value == 2:
                            cell.fill = PatternFill(
                                start_color="FFFF00",
                                end_color="FFFF00",
                                fill_type="solid",
                            )
                        elif value == 1:
                            cell.fill = PatternFill(
                                start_color="FFC000",
                                end_color="FFC000",
                                fill_type="solid",
                            )

            # Update Summary sheet
            ws_summary = wb["Summary"]
            ws_summary.cell(
                row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )

            # Update Performance sheet
            await self._update_performance_sheet(wb)

            wb.save(self.excel_file)

        except Exception as e:
            logger.error(f"Error updating Excel file: {e}")

    async def _update_performance_sheet(self, workbook: openpyxl.Workbook):
        """
        Updates the 'Performance' sheet in the Excel workbook.

        Args:
            workbook (openpyxl.Workbook): The workbook object to update.
        """
        try:
            ws_perf = workbook["Performance"]
            ws_perf.delete_rows(2, ws_perf.max_row + 1)  # Clear existing data

            # Calculate performance by symbol
            symbol_stats = {}
            for signal in self.signal_history:
                symbol = signal.get("Symbol", "")
                if symbol not in symbol_stats:
                    symbol_stats[symbol] = {
                        "total": 0,
                        "confidence_sum": 0,
                        "rr_sum": 0,
                        "last_signal": None,
                    }

                symbol_stats[symbol]["total"] += 1
                symbol_stats[symbol]["confidence_sum"] += signal.get("Confidence", 0)
                symbol_stats[symbol]["rr_sum"] += signal.get("RiskReward", 0)

                signal_time = signal.get("CreatedTime", "")
                if (
                    not symbol_stats[symbol]["last_signal"]
                    or signal_time > symbol_stats[symbol]["last_signal"]
                ):
                    symbol_stats[symbol]["last_signal"] = signal_time

            # Add performance data
            for row, (symbol, stats) in enumerate(symbol_stats.items(), 2):
                avg_confidence = (
                    stats["confidence_sum"] / stats["total"]
                    if stats["total"] > 0
                    else 0
                )
                avg_rr = stats["rr_sum"] / stats["total"] if stats["total"] > 0 else 0

                data = [
                    symbol,
                    stats["total"],
                    0,  # Win rate (would need trade outcomes)
                    round(avg_confidence * 100, 2),
                    round(avg_rr, 2),
                    stats["last_signal"],
                ]

                for col, value in enumerate(data, 1):
                    ws_perf.cell(row=row, column=col, value=value)

        except Exception as e:
            logger.error(f"Error updating performance sheet: {e}")

    async def _update_csv_files(self):
        """Updates the main CSV file with the current active signals."""
        try:
            with open(self.csv_file, "w", newline="") as f:
                writer = csv.writer(f)

                # Headers
                headers = [
                    "ID",
                    "Symbol",
                    "Signal",
                    "Strength",
                    "EntryPrice",
                    "StopLoss",
                    "TakeProfit",
                    "Confidence",
                    "RiskReward",
                    "PositionSize",
                    "MaxRisk",
                    "Timeframe",
                    "Timestamp",
                    "ExpiryTime",
                    "MarketCondition",
                    "TechnicalConfluence",
                    "FundamentalScore",
                    "Status",
                    "CreatedTime",
                    "LastUpdate",
                ]
                writer.writerow(headers)

                # Data
                for signal in self.active_signals.values():
                    row = [
                        signal.get("ID", ""),
                        signal.get("Symbol", ""),
                        signal.get("Signal", ""),
                        signal.get("Strength", ""),
                        signal.get("EntryPrice", 0),
                        signal.get("StopLoss", 0),
                        signal.get("TakeProfit", 0),
                        signal.get("Confidence", 0),
                        signal.get("RiskReward", 0),
                        signal.get("PositionSize", 0),
                        signal.get("MaxRisk", 0),
                        signal.get("Timeframe", ""),
                        signal.get("Timestamp", ""),
                        signal.get("ExpiryTime", ""),
                        signal.get("MarketCondition", ""),
                        signal.get("TechnicalConfluence", 0),
                        signal.get("FundamentalScore", 0),
                        signal.get("Status", ""),
                        signal.get("CreatedTime", ""),
                        signal.get("LastUpdate", ""),
                    ]
                    writer.writerow(row)

        except Exception as e:
            logger.error(f"Error updating CSV files: {e}")

    async def _update_mt4_mt5_files(self):
        """Updates the simplified CSV files intended for MT4 and MT5 EAs."""
        try:
            # MT4 format
            with open(self.mt4_file, "w", newline="") as f:
                writer = csv.writer(f)

                # Simplified headers for MT4 EA
                headers = [
                    "Magic",
                    "Symbol",
                    "Signal",
                    "EntryPrice",
                    "StopLoss",
                    "TakeProfit",
                    "LotSize",
                    "Timestamp",
                ]
                writer.writerow(headers)

                for signal in self.active_signals.values():
                    # Calculate lot size (simplified)
                    lot_size = round(signal.get("PositionSize", 0.01), 2)

                    row = [
                        signal.get("Magic", signal.get("ID", "")),
                        signal.get("Symbol", ""),
                        signal.get("Signal", ""),
                        signal.get("EntryPrice", 0),
                        signal.get("StopLoss", 0),
                        signal.get("TakeProfit", 0),
                        lot_size,
                        signal.get("Timestamp", ""),
                    ]
                    writer.writerow(row)

            # MT5 format (enhanced with more fields)
            with open(self.mt5_file, "w", newline="") as f:
                writer = csv.writer(f)

                headers = [
                    "Magic",
                    "Symbol",
                    "Signal",
                    "EntryPrice",
                    "StopLoss",
                    "TakeProfit",
                    "Volume",
                    "Confidence",
                    "RiskReward",
                    "Expiry",
                    "Comment",
                ]
                writer.writerow(headers)

                for signal in self.active_signals.values():
                    volume = round(signal.get("PositionSize", 0.01), 2)
                    comment = f"GenX_{signal.get('MarketCondition', '')}_{signal.get('Strength', '')}"

                    row = [
                        signal.get("Magic", signal.get("ID", "")),
                        signal.get("Symbol", ""),
                        signal.get("Signal", ""),
                        signal.get("EntryPrice", 0),
                        signal.get("StopLoss", 0),
                        signal.get("TakeProfit", 0),
                        volume,
                        signal.get("Confidence", 0),
                        signal.get("RiskReward", 0),
                        signal.get("ExpiryTime", ""),
                        comment,
                    ]
                    writer.writerow(row)

        except Exception as e:
            logger.error(f"Error updating MT4/MT5 files: {e}")

    async def _update_json_file(self):
        """Updates the JSON file with the current signals and metadata."""
        try:
            data = {
                "signals": list(self.active_signals.values()),
                "last_update": datetime.now().isoformat(),
                "metadata": {
                    "total_signals": len(self.active_signals),
                    "signal_history_count": len(self.signal_history),
                    "update_interval": self.update_interval,
                    "max_signals": self.max_signals,
                },
            }

            with open(self.json_file, "w") as f:
                json.dump(data, f, indent=2, default=str)

        except Exception as e:
            logger.error(f"Error updating JSON file: {e}")

    async def _cleanup_old_signals(self):
        """Removes expired or old signals from the active list."""
        current_time = datetime.now()
        signals_to_remove: List[str] = []

        for signal_id, signal in self.active_signals.items():
            try:
                # Check expiry time
                expiry_str = signal.get("ExpiryTime", "")
                if expiry_str:
                    expiry_time = datetime.strptime(expiry_str, "%Y-%m-%d %H:%M:%S")
                    if current_time > expiry_time:
                        signals_to_remove.append(signal_id)
                        continue

                # Check max age (24 hours)
                created_str = signal.get("CreatedTime", "")
                if created_str:
                    created_time = datetime.strptime(created_str, "%Y-%m-%d %H:%M:%S")
                    if current_time - created_time > timedelta(hours=24):
                        signals_to_remove.append(signal_id)

            except Exception as e:
                logger.error(f"Error checking signal expiry for {signal_id}: {e}")

        # Remove expired signals
        for signal_id in signals_to_remove:
            self.active_signals[signal_id]["Status"] = "EXPIRED"
            del self.active_signals[signal_id]

        # Limit total signals
        if len(self.active_signals) > self.max_signals:
            # Sort by creation time and remove oldest
            sorted_signals = sorted(
                self.active_signals.items(),
                key=lambda x: x[1].get("CreatedTime", ""),
                reverse=True,
            )

            signals_to_keep = dict(sorted_signals[: self.max_signals])
            self.active_signals = signals_to_keep

        if signals_to_remove:
            logger.info(f"Cleaned up {len(signals_to_remove)} expired signals")

    async def _create_backup_if_needed(self):
        """Creates a daily backup of the main Excel signals file."""
        try:
            if not self.last_update:
                return

            backup_date = datetime.now().strftime("%Y-%m-%d")
            backup_file = self.backup_dir / f"signals_backup_{backup_date}.xlsx"

            if not backup_file.exists() and self.excel_file.exists():
                import shutil

                shutil.copy2(self.excel_file, backup_file)
                logger.info(f"Created signals backup: {backup_file}")

        except Exception as e:
            logger.error(f"Error creating backup: {e}")

    async def _load_existing_signals(self):
        """Loads existing active signals from the JSON file on initialization."""
        try:
            if self.json_file.exists():
                with open(self.json_file, "r") as f:
                    data = json.load(f)

                signals = data.get("signals", [])
                for signal in signals:
                    signal_id = signal.get("ID")
                    if signal_id:
                        self.active_signals[signal_id] = signal

                logger.info(f"Loaded {len(signals)} existing signals from JSON file")

        except (json.JSONDecodeError, FileNotFoundError) as e:
            logger.error(f"Error loading existing signals: {e}")

    async def save_final_state(self):
        """Saves the final state of all signal files, intended for graceful shutdown."""
        try:
            await self.update_signals([])  # Trigger an update with no new signals
            logger.info("Saved final spreadsheet state.")
        except Exception as e:
            logger.error(f"Error saving final state: {e}")

    def get_signal_summary(self) -> Dict[str, Any]:
        """
        Gets a summary of the current signal status.

        Returns:
            Dict[str, Any]: A dictionary containing a summary of active signals.
        """
        if not self.active_signals:
            return {"total_signals": 0}

        signals = list(self.active_signals.values())

        # Count by signal type
        buy_signals = sum(1 for s in signals if s.get("Signal") == "BUY")
        sell_signals = sum(1 for s in signals if s.get("Signal") == "SELL")

        # Count by strength
        strength_counts = {}
        for s in signals:
            strength = s.get("Strength", 0)
            strength_counts[strength] = strength_counts.get(strength, 0) + 1

        # Average confidence
        avg_confidence = (
            np.mean([s.get("Confidence", 0) for s in signals]) if signals else 0
        )

        return {
            "total_signals": len(signals),
            "buy_signals": buy_signals,
            "sell_signals": sell_signals,
            "strength_distribution": strength_counts,
            "average_confidence": round(avg_confidence, 3),
            "last_update": self.last_update.isoformat() if self.last_update else None,
            "output_files": {
                "excel": str(self.excel_file),
                "csv": str(self.csv_file),
                "mt4": str(self.mt4_file),
                "mt5": str(self.mt5_file),
                "json": str(self.json_file),
            },
        }
